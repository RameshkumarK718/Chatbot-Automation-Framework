
import openpyxl



class ExcelHandler:

    def __init__(self, file_path: str):

        self.file_path = file_path

        self.workbook = openpyxl.load_workbook(file_path)

        self.sheet = self.workbook.active



    def read_test_cases(self) -> list:

        test_cases = []

        for row in range(2, self.sheet.max_row + 1):

            tc_id = self.sheet.cell(row=row, column=1).value

            if not tc_id:

                continue

            test_cases.append({

                "row": row,

                "id": tc_id,

                "question": self.sheet.cell(row=row, column=2).value,

                "type": self.sheet.cell(row=row, column=3).value or "Direct",

                "context": self.sheet.cell(row=row, column=4).value or "N/A",

                "expected": self.sheet.cell(row=row, column=5).value or ""

            })

        return test_cases



    def update_test_case(self, row: int, actual: str, eval_res: dict):

        self.sheet.cell(row=row, column=6, value=actual)

        

        # Safe extraction to prevent AttributeError when eval_res is an error string

        if isinstance(eval_res, dict):

            relevance = eval_res.get("relevance", "N/A")

            status = eval_res.get("status", "ERROR")

            accuracy = f"{eval_res.get('accuracy_score', 0)}%"

            reason = eval_res.get("evaluation_reason", "N/A")

            remarks = eval_res.get("remarks", "N/A")

        else:

            relevance = "N/A"

            status = "ERROR"

            accuracy = "0%"

            reason = str(eval_res)

            remarks = "Execution Exception"



        self.sheet.cell(row=row, column=7, value=relevance)

        self.sheet.cell(row=row, column=8, value=status)

        self.sheet.cell(row=row, column=9, value=accuracy)

        self.sheet.cell(row=row, column=10, value=reason)

        self.sheet.cell(row=row, column=11, value=remarks)



    def save(self):

        self.workbook.save(self.file_path)

